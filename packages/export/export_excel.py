#!/usr/bin/env python3
"""
Exportação Excel completa — seção 12C.

Gera um .xlsx com 14 abas obrigatórias, cabeçalhos congelados, autofiltro,
formatação de moeda/data e uma aba de Dicionário de Dados. Usa openpyxl
(pré-instalado) — o Excel é uma exportação operacional; o Postgres é a fonte
oficial dos dados (seção 12C, último parágrafo).

Uso: python3 packages/export/export_excel.py [caminho_saida.xlsx]
"""

import sys
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "packages" / "database"))
from db import query  # noqa: E402

TZ = ZoneInfo("America/Sao_Paulo")

HEADER_FILL = PatternFill(start_color="14493D", end_color="14493D", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def write_sheet(wb: Workbook, name: str, headers: list[str], rows: list[list], money_cols: list[int] | None = None):
    ws = wb.create_sheet(name[:31])
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    for row in rows:
        ws.append(row)
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    for i, h in enumerate(headers, start=1):
        max_len = max([len(str(h))] + [len(str(r[i - 1])) for r in rows]) if rows else len(str(h))
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, 10), 60)
    if money_cols:
        for col in money_cols:
            for r in range(2, len(rows) + 2):
                cell = ws.cell(row=r, column=col)
                if isinstance(cell.value, (int, float)):
                    cell.value = cell.value / 100
                    cell.number_format = '"R$" #,##0.00'
    return ws


def build_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    # 1. Leads
    leads = query(
        "SELECT c.id, c.trade_name, c.legal_name, c.cnpj, c.city, c.state, seg.name AS segmento, "
        "c.status, c.website, c.instagram, c.created_at, c.last_validated_at "
        "FROM companies c LEFT JOIN segments seg ON seg.id=c.segment_id ORDER BY c.id"
    )
    write_sheet(wb, "Leads",
                ["ID", "Nome fantasia", "Razão social", "CNPJ", "Cidade", "UF", "Segmento", "Status",
                 "Site", "Instagram", "Criado em", "Última validação"],
                [[r["id"], r["trade_name"], r["legal_name"], r["cnpj"], r["city"], r["state"], r["segmento"],
                  r["status"], r["website"], r["instagram"], r["created_at"], r["last_validated_at"]] for r in leads])

    # 2. Contatos
    contacts = query(
        "SELECT ct.id, c.id AS company_id, c.trade_name, ct.name, ct.role, ct.email, ct.phone, "
        "ct.whatsapp, ct.confidence FROM contacts ct JOIN companies c ON c.id=ct.company_id ORDER BY ct.id"
    )
    write_sheet(wb, "Contatos",
                ["ID", "Empresa ID", "Empresa", "Nome", "Cargo", "E-mail", "Telefone", "WhatsApp", "Confiança"],
                [[r["id"], r["company_id"], r["trade_name"], r["name"], r["role"], r["email"], r["phone"],
                  r["whatsapp"], r["confidence"]] for r in contacts])

    # 3. Auditoria Marketplace
    audit = query(
        "SELECT mp.id, c.trade_name, mp.channel, mp.present, mp.listings_count, mp.price_alignment, "
        "mp.confidence, mp.evidence_url, mp.checked_at FROM marketplace_presences mp "
        "JOIN companies c ON c.id=mp.company_id ORDER BY mp.id"
    )
    write_sheet(wb, "Auditoria Marketplace",
                ["ID", "Empresa", "Canal", "Presente?", "Anúncios", "Alinhamento de preço", "Confiança",
                 "Evidência (URL)", "Verificado em"],
                [[r["id"], r["trade_name"], r["channel"], "Sim" if r["present"] == "t" else "Não",
                  r["listings_count"], r["price_alignment"], r["confidence"], r["evidence_url"], r["checked_at"]]
                 for r in audit])

    # 4. Score
    scores = query(
        "SELECT ls.id, c.trade_name, ls.product_fit, ls.marketplace_gap, ls.business_structure, "
        "ls.catalog_quality, ls.investment_signals, ls.contactability, ls.problem_clarity, ls.total, "
        "ls.potential, ls.confidence, ls.rationale, ls.generated_at "
        "FROM lead_scores ls JOIN companies c ON c.id=ls.company_id ORDER BY ls.total DESC"
    )
    write_sheet(wb, "Score",
                ["ID", "Empresa", "Product fit", "Marketplace gap", "Estrutura", "Catálogo",
                 "Sinais invest.", "Contactabilidade", "Clareza problema", "Total", "Potencial",
                 "Confiança", "Justificativa", "Gerado em"],
                [[r["id"], r["trade_name"], r["product_fit"], r["marketplace_gap"], r["business_structure"],
                  r["catalog_quality"], r["investment_signals"], r["contactability"], r["problem_clarity"],
                  r["total"], r["potential"], r["confidence"], r["rationale"], r["generated_at"]] for r in scores])

    # 5. Abordagens
    attempts = query(
        "SELECT oa.id, c.trade_name, oa.channel, oa.status, oa.attempt_number, oa.human_confirmed, "
        "oa.sent_at, om.step, om.text FROM outreach_attempts oa JOIN companies c ON c.id=oa.company_id "
        "LEFT JOIN outreach_messages om ON om.id = oa.message_id ORDER BY oa.id"
    )
    write_sheet(wb, "Abordagens",
                ["ID", "Empresa", "Canal", "Status", "Nº tentativa", "Confirmado por humano?", "Enviado em",
                 "Etapa", "Texto"],
                [[r["id"], r["trade_name"], r["channel"], r["status"], r["attempt_number"],
                  "Sim" if r["human_confirmed"] == "t" else "Não", r["sent_at"], r["step"], r["text"]]
                 for r in attempts])

    # 6. Follow-ups (tarefas)
    tasks = query(
        "SELECT t.id, c.trade_name, t.title, t.due_at, t.status FROM tasks t "
        "LEFT JOIN companies c ON c.id=t.company_id ORDER BY t.due_at"
    )
    write_sheet(wb, "Follow-ups",
                ["ID", "Empresa", "Tarefa", "Vencimento", "Status"],
                [[r["id"], r["trade_name"], r["title"], r["due_at"], r["status"]] for r in tasks])

    # 7. Agenda
    agenda = query(
        "SELECT ce.id, c.trade_name, ce.title, ce.starts_at, ce.ends_at, ce.status FROM calendar_events ce "
        "LEFT JOIN companies c ON c.id=ce.company_id ORDER BY ce.starts_at"
    )
    write_sheet(wb, "Agenda",
                ["ID", "Empresa", "Título", "Início", "Fim", "Status"],
                [[r["id"], r["trade_name"], r["title"], r["starts_at"], r["ends_at"], r["status"]] for r in agenda])

    # 8. Reuniões
    meetings = query(
        "SELECT m.id, c.trade_name, m.type, m.attended FROM meetings m JOIN companies c ON c.id=m.company_id ORDER BY m.id"
    )
    write_sheet(wb, "Reuniões",
                ["ID", "Empresa", "Tipo", "Compareceu?"],
                [[r["id"], r["trade_name"], r["type"], r["attended"]] for r in meetings])

    # 9. Propostas
    proposals = query(
        "SELECT p.id, c.trade_name, sp.name AS oferta, p.status, p.valid_until, pv.final_price_cents, "
        "pv.payment_terms FROM proposals p JOIN companies c ON c.id=p.company_id "
        "JOIN service_packages sp ON sp.id=p.offer_id "
        "LEFT JOIN LATERAL (SELECT * FROM proposal_versions WHERE proposal_id=p.id ORDER BY version DESC LIMIT 1) pv ON TRUE "
        "ORDER BY p.id"
    )
    write_sheet(wb, "Propostas",
                ["ID", "Empresa", "Oferta", "Status", "Válida até", "Valor final (R$)", "Condição de pagamento"],
                [[r["id"], r["trade_name"], r["oferta"], r["status"], r["valid_until"],
                  int(r["final_price_cents"]) if r["final_price_cents"] else None, r["payment_terms"]]
                 for r in proposals],
                money_cols=[6])

    # 10. Contratos
    contracts = query(
        "SELECT ct.id, c.trade_name, ct.status, ct.sent_at, ct.signed_at FROM contracts ct "
        "JOIN companies c ON c.id=ct.company_id ORDER BY ct.id"
    )
    write_sheet(wb, "Contratos",
                ["ID", "Empresa", "Status", "Enviado em", "Assinado em"],
                [[r["id"], r["trade_name"], r["status"], r["sent_at"], r["signed_at"]] for r in contracts])

    # 11. Pagamentos
    payments = query(
        "SELECT i.id, c.trade_name, i.seq, i.amount_cents, i.due_date, i.status, i.paid_at "
        "FROM installments i JOIN payments pay ON pay.id=i.payment_id JOIN contracts ct ON ct.id=pay.contract_id "
        "JOIN companies c ON c.id=ct.company_id ORDER BY i.due_date"
    )
    write_sheet(wb, "Pagamentos",
                ["ID", "Empresa", "Parcela", "Valor (R$)", "Vencimento", "Status", "Pago em"],
                [[r["id"], r["trade_name"], r["seq"], int(r["amount_cents"]), r["due_date"], r["status"], r["paid_at"]]
                 for r in payments],
                money_cols=[4])

    # 12. Dashboard (indicadores)
    total_pipeline = query(
        "SELECT COALESCE(SUM(value_proposed_cents),0) AS v FROM deals d JOIN pipeline_stages ps ON ps.id=d.stage_id "
        "WHERE ps.code NOT IN ('perdido','bloqueado','ganho_recebido')"
    )[0]["v"]
    total_received = query("SELECT COALESCE(SUM(amount_cents),0) AS v FROM revenue_events")[0]["v"]
    n_companies = query("SELECT count(*) AS n FROM companies")[0]["n"]
    n_qualified = query("SELECT count(*) AS n FROM lead_scores WHERE potential IN ('A','B')")[0]["n"]
    n_contacted = query("SELECT count(*) AS n FROM outreach_attempts")[0]["n"]
    n_replies = query("SELECT count(*) AS n FROM reply_classifications")[0]["n"]
    n_meetings = query("SELECT count(*) AS n FROM meetings")[0]["n"]
    n_proposals = query("SELECT count(*) AS n FROM proposals")[0]["n"]
    n_won = query("SELECT count(*) AS n FROM deals d JOIN pipeline_stages ps ON ps.id=d.stage_id WHERE ps.code='ganho_recebido'")[0]["n"]
    dash_rows = [
        ["Empresas pesquisadas", int(n_companies)],
        ["Leads qualificados (A/B)", int(n_qualified)],
        ["Contatos realizados", int(n_contacted)],
        ["Respostas recebidas", int(n_replies)],
        ["Reuniões", int(n_meetings)],
        ["Propostas emitidas", int(n_proposals)],
        ["Negócios ganhos e recebidos", int(n_won)],
        ["Pipeline em aberto (R$)", int(total_pipeline) / 100],
        ["Dinheiro efetivamente recebido (R$)", int(total_received) / 100],
    ]
    ws_dash = write_sheet(wb, "Dashboard", ["Indicador", "Valor"], dash_rows)
    ws_dash["B8"].number_format = '"R$" #,##0.00'
    ws_dash["B9"].number_format = '"R$" #,##0.00'
    ws_dash.append(["Total (soma dos itens numéricos de contagem)", "=SUM(B2:B7)"])

    # 13. Motivos de Perda
    lost = query(
        "SELECT d.id, c.trade_name, d.loss_reason FROM deals d JOIN companies c ON c.id=d.company_id "
        "JOIN pipeline_stages ps ON ps.id=d.stage_id WHERE ps.code='perdido'"
    )
    write_sheet(wb, "Motivos de Perda",
                ["Deal ID", "Empresa", "Motivo"],
                [[r["id"], r["trade_name"], r["loss_reason"]] for r in lost])

    # 14. Dicionário de Dados
    dictionary = [
        ["Leads", "status", "Estágio operacional da empresa no funil (ver aba Dicionário / seção 12D)"],
        ["Score", "total", "Soma das 7 componentes de 0 a 100 — seção 6.3 do prompt-mestre"],
        ["Score", "potential", "A (80-100), B (65-79), NUTRIR (50-64), NAO_ABORDAR (<50)"],
        ["Abordagens", "human_confirmed", "TRUE somente após clique humano de envio (seção 15, Modo 1 — Assistido)"],
        ["Pagamentos", "status", "PENDING | RECEIVED | LATE | CANCELLED"],
        ["Dashboard", "Dinheiro efetivamente recebido", "Só conta após confirmação do 1º pagamento — seção 3.5 / 12H"],
        ["Geral", "Todas as empresas e valores", "Dados fictícios de demonstração — seção 16, nunca dados reais em teste"],
    ]
    write_sheet(wb, "Dicionário de Dados", ["Aba", "Campo", "Descrição"], dictionary)

    return wb


def main():
    out_path = sys.argv[1] if len(sys.argv) > 1 else str(ROOT / "exports" / "marketplace_growth_engine_export.xlsx")
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook()
    now = datetime.now(TZ)
    meta_ws = wb.create_sheet("Metadados", 0)
    meta_ws.append(["Exportado em", now.strftime("%Y-%m-%d %H:%M:%S %Z")])
    meta_ws.append(["Fuso horário", "America/Sao_Paulo"])
    meta_ws.append(["Fonte oficial dos dados", "PostgreSQL (packages/database) — este Excel é apenas uma exportação"])
    meta_ws.column_dimensions["A"].width = 30
    meta_ws.column_dimensions["B"].width = 50
    wb.save(out_path)
    print(f"Exportado para {out_path}")


if __name__ == "__main__":
    main()
