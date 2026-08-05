#!/usr/bin/env python3
"""
Gera a planilha manual de validação do piloto (passo #1).

Uso:
    python3 scripts/gerar_planilha_validacao.py [saida.xlsx]

Sem depender do painel/Postgres: é a ferramenta de acompanhamento que o Iago
preenche na mão durante a validação da oferta. Ver docs/validacao-oferta.md.
Regras: nunca dado real em modelo/seed — as linhas de exemplo são fictícias
e marcadas como tal (seção 16).
"""
import sys
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

AZUL = "1F3A5F"
CINZA = "E8ECF1"
VERDE = "2E7D32"

HEADER_FILL = PatternFill("solid", fgColor=AZUL)
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SUB_FILL = PatternFill("solid", fgColor=CINZA)
BOLD = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Side(style="thin", color="C0C8D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _cabecalho(ws, colunas, larguras):
    for i, (titulo, larg) in enumerate(zip(colunas, larguras), start=1):
        c = ws.cell(row=1, column=i, value=titulo)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        c.border = BORDER
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = larg
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def aba_instrucoes(wb):
    ws = wb.active
    ws.title = "Instruções"
    linhas = [
        ("MÁQUINA DE VENDAS — VALIDAÇÃO DA OFERTA (PASSO #1)", True),
        ("", False),
        ("Objetivo: descobrir se a oferta vende antes de automatizar.", False),
        ("Uma campanha, um segmento: Casa / Móveis / Decoração.", False),
        ("Playbook completo: docs/validacao-oferta.md", False),
        ("", False),
        ("Metas do piloto:", True),
        ("  • 100 empresas pesquisadas", False),
        ("  • 40 leads com score >= 65", False),
        ("  • 20 contatos aprovados manualmente", False),
        ("", False),
        ("Como usar as abas:", True),
        ("  1. Leads — uma linha por empresa. Pontue (score manual), classifique A/B/NUTRIR/NÃO.", False),
        ("  2. Só aprove abordagem para A e B. Registre cada envio da cadência e a resposta.", False),
        ("  3. Métricas — atualize os números ao longo da campanha (funil).", False),
        ("  4. Motivos de perda — pós-mortem de cada negócio perdido.", False),
        ("  5. Decisão — go/no-go ao fim da campanha (escalar / ajustar / trocar / parar).", False),
        ("", False),
        ("Compliance (obrigatório):", True),
        ("  • Cada envio de WhatsApp é clique humano. Só horário comercial.", False),
        ("  • Respeitar opt-out na hora. Sem dado sensível. Não inventar CNPJ/faturamento.", False),
        ("  • WhatsApp dedicado ao comercial (não o pessoal).", False),
        ("", False),
        ("As linhas de exemplo são de empresas FICTÍCIAS. Apague antes de usar.", True),
    ]
    ws.column_dimensions["A"].width = 95
    for i, (texto, negrito) in enumerate(linhas, start=1):
        c = ws.cell(row=i, column=1, value=texto)
        if i == 1:
            c.font = Font(bold=True, size=14, color=AZUL)
        elif negrito:
            c.font = Font(bold=True, color=VERDE)
    return ws


def aba_leads(wb):
    ws = wb.create_sheet("Leads")
    colunas = [
        "Empresa", "Cidade/UF", "Site/Instagram", "Contato (nome)", "Telefone/e-mail",
        "SKUs (aprox.)", "Anos ativo", "Marca própria?", "Presença marketplace (gap)",
        "Score (0-100)", "Classe (A/B/NUTRIR/NÃO)", "Aprovado p/ abordar?",
        "Data contato inicial", "Follow-up 1", "Follow-up 2", "Encerramento",
        "Respondeu?", "Reunião marcada?", "Proposta enviada?", "Oferta / Valor",
        "Status", "Opt-out?", "Observações",
    ]
    larguras = [22, 12, 24, 16, 20, 11, 10, 13, 22, 12, 16, 15, 15, 12, 12, 13, 12, 14, 14, 18, 16, 10, 30]
    _cabecalho(ws, colunas, larguras)

    # validações de lista
    dv_classe = DataValidation(type="list", formula1='"A,B,NUTRIR,NÃO ABORDAR"', allow_blank=True)
    dv_simnao = DataValidation(type="list", formula1='"Sim,Não"', allow_blank=True)
    dv_status = DataValidation(
        type="list",
        formula1='"Novo,Aprovado,Contatado,Em conversa,Reunião marcada,Proposta enviada,Ganho,Perdido,Nutrir"',
        allow_blank=True,
    )
    for dv in (dv_classe, dv_simnao, dv_status):
        ws.add_data_validation(dv)
    dv_classe.add("K2:K500")
    dv_status.add("U2:U500")
    for col in ("H", "L", "Q", "R", "S", "V"):
        dv_simnao.add(f"{col}2:{col}500")

    exemplos = [
        ["Casa Bela Utilidades (FICTÍCIA)", "Curitiba/PR", "@casabela", "Marina", "comercial@…",
         120, 6, "Sim", "Ausente no ML; revendedores na Amazon", 82, "A", "Sim",
         str(date.today()), "", "", "", "Sim", "Sim", "Não", "Implantação 90 / R$20.000",
         "Reunião marcada", "Não", "Marca forte em cama/mesa/banho"],
        ["Móveis Rio Verde (FICTÍCIA)", "Goiânia/GO", "site próprio", "Paulo", "(62)…",
         60, 4, "Sim", "Execução fraca no ML", 68, "B", "Sim",
         str(date.today()), "", "", "", "Não", "Não", "Não", "Diagnóstico / R$2.500",
         "Contatado", "Não", "Aguardando resposta do 1º contato"],
    ]
    for r, linha in enumerate(exemplos, start=2):
        for c, valor in enumerate(linha, start=1):
            cell = ws.cell(row=r, column=c, value=valor)
            cell.border = BORDER
            cell.alignment = WRAP
            cell.font = Font(italic=True, color="8A94A0")
    return ws


def aba_metricas(wb):
    ws = wb.create_sheet("Métricas")
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 40
    ws.merge_cells("A1:D1")
    t = ws.cell(row=1, column=1, value="FUNIL DA CAMPANHA — CASA / MÓVEIS / DECORAÇÃO")
    t.fill = HEADER_FILL
    t.font = HEADER_FONT
    t.alignment = CENTER
    ws.row_dimensions[1].height = 26
    for col, titulo in zip("ABCD", ["Etapa", "Meta", "Real", "Taxa de conversão"]):
        c = ws[f"{col}2"]
        c.value = titulo
        c.fill = SUB_FILL
        c.font = BOLD
        c.border = BORDER
    funil = [
        ("Empresas pesquisadas", 100, ""),
        ("Leads com score >= 65", 40, ""),
        ("Contatos aprovados", 20, ""),
        ("Mensagens enviadas", "", ""),
        ("Respostas recebidas", "", ""),
        ("Respostas positivas", "", ""),
        ("Reuniões de diagnóstico", "", ""),
        ("Propostas enviadas", "", ""),
        ("Fechamentos (entrada paga)", "", ""),
        ("Opt-outs", "", ""),
    ]
    for i, (etapa, meta, real) in enumerate(funil, start=3):
        ws[f"A{i}"] = etapa
        ws[f"B{i}"] = meta
        ws[f"C{i}"] = real
        # taxa = etapa atual / etapa anterior
        if i > 3:
            ws[f"D{i}"] = f'=IFERROR(C{i}/C{i-1},"")'
            ws[f"D{i}"].number_format = "0.0%"
        for col in "ABCD":
            ws[f"{col}{i}"].border = BORDER
    return ws


def aba_perdas(wb):
    ws = wb.create_sheet("Motivos de perda")
    colunas = ["Empresa", "Classe", "Etapa em que perdeu", "Motivo principal",
               "Objeção central", "Aprendizado", "Retomar? Quando"]
    larguras = [24, 10, 22, 32, 24, 34, 20]
    _cabecalho(ws, colunas, larguras)
    dica = ["(ex.: Casa Bela — FICTÍCIA)", "A", "Após diagnóstico", "Preço acima do orçamento no semestre",
            "Preço", "Fundador topou valor menor; testar oferta Diagnóstico como entrada", "Q4"]
    for c, v in enumerate(dica, start=1):
        cell = ws.cell(row=2, column=c, value=v)
        cell.font = Font(italic=True, color="8A94A0")
        cell.alignment = WRAP
        cell.border = BORDER
    return ws


def aba_decisao(wb):
    ws = wb.create_sheet("Decisão")
    ws.column_dimensions["A"].width = 100
    blocos = [
        ("DECISÃO GO / NO-GO AO FIM DA CAMPANHA", "titulo"),
        ("", None),
        ("Preencha depois das 20 abordagens aprovadas:", "bold"),
        ("", None),
        ("As 5 perguntas do piloto (responda com dados reais):", "bold"),
        ("1. Qual segmento mais respondeu?", None),
        ("2. Qual mensagem gerou conversa?", None),
        ("3. Qual perfil fechou (faturamento / SKUs / decisor)?", None),
        ("4. Qual foi a objeção principal?", None),
        ("5. Quantos projetos o Iago consegue tocar bem ao mesmo tempo?", None),
        ("", None),
        ("Decisão (escolha uma):", "bold"),
        ("[ ] ESCALAR — >=3 reuniões e >=1 proposta com fit -> 2ª campanha / iniciar automação", None),
        ("[ ] AJUSTAR — houve resposta mas nenhuma reunião -> nova mensagem/ICP, mesmo segmento", None),
        ("[ ] TROCAR DE SEGMENTO — resposta < 5% e zero conversa -> testar Moda (segmento #2)", None),
        ("[ ] PARAR E REAVALIAR — reuniões travam sempre no mesmo ponto -> revisar oferta/preço", None),
        ("", None),
        ("Justificativa da decisão:", "bold"),
        ("", None),
    ]
    for i, (texto, estilo) in enumerate(blocos, start=1):
        c = ws.cell(row=i, column=1, value=texto)
        if estilo == "titulo":
            c.font = Font(bold=True, size=14, color=AZUL)
        elif estilo == "bold":
            c.font = Font(bold=True, color=VERDE)
    return ws


def main():
    saida = sys.argv[1] if len(sys.argv) > 1 else "exports/validacao-piloto.xlsx"
    wb = Workbook()
    aba_instrucoes(wb)
    aba_leads(wb)
    aba_metricas(wb)
    aba_perdas(wb)
    aba_decisao(wb)
    wb.save(saida)
    print(f"Planilha de validação gerada em: {saida}")


if __name__ == "__main__":
    main()
